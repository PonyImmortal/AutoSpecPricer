from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


def format_excel_from_dataframe(workbook, dataframe, file_name):
    ws = workbook.active

    EQUIPMENT_SPECIFICATION_COLUMN = 1
    SIMILAR_EQUIPMENT_COLUMN = 5
    SIZE_SPECIFICATION_COLUMN = 2
    SIZE_PRICE_COLUMN = 7
    PRICE_COLUMN = 8
    NOT_FOUND_MESSAGE = "Совпадений не найдено"

    thin_border = Border(left=Side(style='thin', color="D3D3D3"),
                         right=Side(style='thin', color="D3D3D3"),
                         top=Side(style='thin', color="D3D3D3"),
                         bottom=Side(style='thin', color="D3D3D3"))

    for r_idx, row in enumerate(dataframe_to_rows(dataframe, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border

            if c_idx == SIZE_SPECIFICATION_COLUMN or c_idx == SIZE_PRICE_COLUMN or c_idx == PRICE_COLUMN:
                cell.font = Font(bold=True)

            if c_idx == SIMILAR_EQUIPMENT_COLUMN:
                if value == NOT_FOUND_MESSAGE:
                    cell.font = Font(color="FF0000")
                    ws.cell(row=r_idx, column=EQUIPMENT_SPECIFICATION_COLUMN).font = Font(color="FF0000")

            if c_idx in [EQUIPMENT_SPECIFICATION_COLUMN, SIZE_SPECIFICATION_COLUMN]:
                cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
            elif c_idx in [SIMILAR_EQUIPMENT_COLUMN, SIZE_PRICE_COLUMN, PRICE_COLUMN]:
                cell.fill = PatternFill(start_color="F5FFEE", end_color="F5FFEE", fill_type="solid")

    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            cell_value = cell.value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    workbook.save(file_name)

